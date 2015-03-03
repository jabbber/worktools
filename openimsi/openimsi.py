#!/usr/bin/env python2
# -*- coding: utf-8 -*-
import os,sys
import traceback
import re
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import openpyxl.style
from bs4 import BeautifulSoup

class Tables():
    def __init__(self):
        self.titles = []
        self.tables = []
        self.hostlist = []
        self.blacklist = []
        self.project_find = None
    def clear(self):
        """clear the data have been loaded.
        """
        self.titles = []
        self.tables = []
    def load(self,filename):
        """load data from given file path,the file can be .xlsx or .htm(l)
        """
        name, ext = os.path.splitext(filename)
        if ext == ".xlsx":
            try:
                self.load_xlsx(filename)
            except:
                print traceback.print_exc()
                return False
        elif ext in ('.htm','.html'):
            try:
                self.load_html(filename)
            except:
                print traceback.print_exc()
                return False
        else:
            print "error: can load %s!"%filename
            return False
        return True
    def load_html(self,html_file):
        html = file(os.path.realpath(html_file),'r').read().decode('utf-8')
        soup = BeautifulSoup(html)
        if soup.html:
            tables = soup.html.body.find_all('table',recursive=False)
        else:
            tables = soup.find_all('table',recursive=False)
        for table in tables:
            try:
                title = table.find_previous('p').get_text().encode('utf-8')
            except:
                title = ''
            self.titles.append(title)
            self.tables.append([])
            for row in table.find_all('tr',recursive=False):
                line = []
                for val in row.find_all('td',recursive=False):
                    value = val.get_text()
                    if type(value) == unicode:
                        value = value.encode('utf-8')
                    line.append(value)
                if self.__black_filter(line):
                    pass
                elif self.__host_filter(line):
                    self.tables[-1].append(line)
                else:
                    pass
    def load_xlsx(self,xlsx_file):
        wb = load_workbook(filename = xlsx_file)
        sheet_ranges = wb.get_active_sheet()
        table = 0
        row = 0
        row_empty = 0
        row_max_empty = 100
        end = False
        th = None
        while not end:
            col = []
            col_empty = 0
            col_max_empty = 3
            col_end = False
            while not col_end:
                value = sheet_ranges.cell(row=row,column=len(col)).value
                if value:
                    if type(value) == unicode:
                        value = value.encode('utf-8')
                    col_empty = 0
                else:
                    value = ''
                    col_empty += 1
                col.append(value)
                if col_empty >= col_max_empty:
                    col = col[:-col_max_empty]
                    col_end = True
            if len(col) == 1:
                if not th:
                    th = col[0]
                    worktable = []
                else:
                    if th == "" and worktable == []:
                        pass
                    else:
                        self.titles.append(th)
                        self.tables.append(worktable)
                    th = col[0]
                    worktable = []
            elif len(col) > 1:
                if not th:
                    th = ""
                    worktable = []
                if self.__black_filter(col):
                    pass
                elif self.__host_filter(col):
                    worktable.append(col)
                else:
                    pass
                row_empty = 0
            else:
                row_empty += 1
            if row_empty >= row_max_empty:
                end = True
                self.titles.append(th)
                self.tables.append(worktable)
            row += 1
        return 0
    def __load_xlsx_list(self,xlsx_file):
        wb = load_workbook(filename = xlsx_file)
        sheet_ranges = wb.get_active_sheet()
        end = False
        row = 0
        keylist = []
        while not end:
            value = sheet_ranges.cell(row=row,column=0).value
            if type(value) == unicode:
                keylist.append(value.encode('utf-8'))
            else:
                end = True
            row += 1
        return keylist
    def load_list(self,list_file):
        """load a list from a text or xlsx file,return python list type.
        """
        if not list_file:
            keylist = []
        elif list_file[-5:] == '.xlsx':
            keylist = self.__load_xlsx_list(list_file)
        else:
            with open(list_file) as list_f:
                keylist = [line.strip() for line in list_f if line.strip() != '']
        return keylist
    def __filter(self,row,key_list):
        for value in row:
            for key in key_list:
                if type(key) == str and type(value) == str:
                    if value.find(key) >= 0 or value.find(key.upper()) >=0:
                        return True
        return False
    def load_project_list(self,list_file):
        self.__project_find = '''
            HQs(\w{3})?                         #Host type
            [_|-]?                              #split sign
            (''' + '|'.join(self.load_list(list_file)) +''')      #project name
            [_|-]?                              #split sign
            (\d|\w\d{0,2}|\w{3})                #number
        '''
    def __host_search(self,string):
        if re.search(self.__project_find, string, re.VERBOSE):
            return True
        else:
            return False
    def __host_filter(self,row):
        if self.project_find:
            return self.__host_search(row)

        key_list = self.hostlist
        if key_list == []:
            return True
        if '序号' == row[0]:
            return True
        else:
            return self.__filter(row,key_list)
    def __black_filter(self,row):
        key_list = self.blacklist
        if key_list == []:
            return False
        else:
            return self.__filter(row,key_list)
    def __xlsx_setstyle(self,style,style_string):
        argv = style_string.split(':')
        n = 0
        fields = {'font':('name','size','bold','italic','superscript','subscript','underline','strikethrough','color'),
                'fill':('fill_type','rotation','start_color','end_color'),
                'borders':('left','right','top','bottom','diagonal','diagonal_direction','all_borders','outline','inside','vertical','horizontal'),
                'alignment':('horizontal','vertical','text_rotation','wrap_text','shrink_to_fit','indent'),
                'number_format':('_format_code','_format_index'),
                'protection':('locked','hidden')
                }
        for obj in ('font','fill','borders','alignment','number_format','protection'):
            for val in fields[obj]:
                if obj == 'borders' and val != 'diagonal_direction':
                    for val_1 in ('border_style','color'):
                        exec "style.%s.%s.%s = eval(argv[n])"%(obj,val,val_1)
                        n += 1
                else:
                    exec "style.%s.%s = eval(argv[n])"%(obj,val)
                    n += 1
        return style
    def get_html(self,titles = 'all'):
        """return the tables you given titles use html table format,default is all tables.
        """
        tables = self.tables
        output = ''
        if titles == 'all':
            for i, title in enumerate(self.titles):
                output += ('<p>%s</p>\n'%title)
                output += table2html(tables[i])
                output += '<br/>\n'
        else:
            for title in titles.split(','):
#                try:
#                    title = title.decode('utf-8')
#                except:
#                    title = title.decode('gb18030')
                output += ('%s<br/>\n'%title)
                n = 0
                finded = []
                for value in self.titles:
                    if value == title:
                        finded.append(n)
                    n += 1
                for num in finded:
                    output += table2html(tables[num])
                    output += '<br/>\n'
        return output
    def get_xlsx(self,dest_filename):
        """create a .xlsx file have all data have been loaded to you given path and filename.
        """
        tables = self.tables
        wb = Workbook()
        ws = wb.worksheets[0]
        row_num = 0
        column_widths = []
        for title_num, title in enumerate(self.titles):
#            ws.title = title
            row_num += 1
            cell = ws.cell('%s%s'%(get_column_letter(1), row_num))
            cell.value = title
            cell.style.font.size = 13
            row_num += 1
            for row in tables[title_num]:
                row_num += 1
                col_num = 0
                th = False
                for i, value in enumerate(row):
                    # count column width
                    if type(value) == str:
                        l = len(value.decode('utf-8'))
                        if l == len(value):
                            pass
                        else:
                            l = l*2
                    elif type(value) == int:
                        l = len(str(value))
                    else:
                        l = len(value)
                    if len(column_widths) > i:
                        if l > column_widths[i]:
                                column_widths[i] = l
                    else:
                        column_widths.append(l)

                    col_num += 1
                    cell = ws.cell('%s%s'%(get_column_letter(col_num), row_num))
                    cell.value = value
                    cell.style.font.size = 10
                    for key in ('left','right','top','bottom'):
                        exec "cell.style.borders.%s.border_style = 'medium'"%key
                    if value == '序号':
                        th = True
                    if th:
#                        self.__xlsx_setstyle(cell.style,self.XLSXSTYLE_title)
                        cell.style.font.bold = True
                    else:
                        cell.style.alignment.wrap_text = True
            row_num += 1
        for i, column_width in enumerate(column_widths):
            if column_width < 6:
                ws.column_dimensions[get_column_letter(i+1)].width = 6
            elif column_width > 40:
                ws.column_dimensions[get_column_letter(i+1)].width = 40
            else:
                ws.column_dimensions[get_column_letter(i+1)].width = column_width

#            ws = wb.create_sheet()
#        wb.worksheets.pop()

        wb.save(filename = dest_filename)

def table2html(table):
    column_widths = []
    html = '<table>\n'
    for row in table:
        if row[0] == '序号':
            html += '  <tr>\n'
            for i, val in enumerate(row):
                html += "<th>%s</th>"%val
            html += '  </tr>\n'
        else:
            html += '  <tr>\n'
            for val in row:
                html += "<td>%s</td>"%val
            html += '  </tr>\n'
    html += '</table>'
    return html

def get_html(filename,titles = 'all',host_file = None,blacklist = None):
    t = Tables()
    if host_file:
        t.hostlist = t.load_list(host_file)
    if blacklist:
        t.blacklist = t.load_list(blacklist)
    t.load(filename)
    return t.get_html(titles)

if __name__ == "__main__":
    if len(sys.argv) == 5:
        print get_html(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4])
    elif len(sys.argv) == 4:
        print get_html(sys.argv[1],sys.argv[2],sys.argv[3])
    elif len(sys.argv) == 3:
        print get_html(sys.argv[1],sys.argv[2])
    else:
        print """Usage:
    readxlsx.py <FILE> all/title1[,title2] [HOST_FILE] [Blacklist_FILE]

FILE can be an unix text file or xlsx file.
        """
