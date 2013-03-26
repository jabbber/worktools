#!/usr/bin/env python2
# -*- coding: utf-8 -*-
import os,sys
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import openpyxl.style
from bs4 import BeautifulSoup

class Tables():
    XLSXSTYLE = "'Calibri':10:False:False:False:False:'none':False:'FF000000':'none':0:'FFFFFFFF':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'none':'FF000000':0:'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'general':'center':0:True:False:0:'General':0:'inherit':'inherit'"
    XLSXSTYLE_title = "'Calibri':10:True:False:False:False:'none':False:'FF0000FF':'none':0:'FFFFFFFF':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'none':'FF000000':0:'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'general':'center':0:False:False:0:'General':0:'inherit':'inherit'"
    def __init__(self):
        self.titles = []
        self.tables = {}
        self.hostlist = []
        self.blacklist = []
    def load(self,filename):
        ext = filename[filename.find('.')+1:]
        if ext == "xlsx":
            self.load_xlsx(filename)
        elif ext in ('htm','html'):
            self.load_html(filename)
        else:
            print "error: can load '%s' file!"%ext
    def load_html(self,html_file):
        html = file(os.path.realpath(html_file),'r').read().decode('utf-8')
        soup = BeautifulSoup(html)
        tables = soup.find_all('table')
        for table in tables:
            title = table.find_previous('p').get_text()
            self.titles.append(title)
            self.tables[title] = []
            for row in table.find_all('tr'):
                line = []
                for val in row.find_all('td'):
                    value = val.get_text()
                    if type(value) == unicode:
                        value = value.encode('utf-8')
                    line.append(value)
                if self.__black_filter(line):
                    pass
                elif self.__host_filter(line):
                    self.tables[title].append(line)
                else:
                    pass
    def load_xlsx(self,xlsx_file):
        wb = load_workbook(filename = xlsx_file)
        sheet_ranges = wb.get_active_sheet()
        table = 0
        row = 0
        row_empty = 0
        row_max_empty = 100
        end = False
        th = None
        while not end:
            col = []
            col_empty = 0
            col_max_empty = 3
            col_end = False
            while not col_end:
                value = sheet_ranges.cell(row=row,column=len(col)).value
                if value:
                    col_empty = 0
                else:
                    value = ''
                    col_empty += 1
                col.append(value)
                if col_empty >= col_max_empty:
                    col = col[:-col_max_empty]
                    col_end = True
            if len(col) == 1:
                if not th:
                    th = col[0]
                    worktable = []
                else:
                    if th == "" and worktable == []:
                        pass
                    else:
                        self.titles.append(th)
                        self.tables[th] = worktable
                    th = col[0]
                    worktable = []
            elif len(col) > 1:
                if not th:
                    th = ""
                    worktable = []
                worktable.append(col)
                row_empty = 0
            else:
                row_empty += 1
            if row_empty >= row_max_empty:
                end = True
                self.titles.append(th)
                self.tables[th] = worktable
            row += 1
        return 0
    def __load_xlsx_list(self,xlsx_file):
        wb = load_workbook(filename = xlsx_file)
        sheet_ranges = wb.get_active_sheet()
        end = False
        row = 0
        keylist = []
        while not end:
            value = sheet_ranges.cell(row=row,column=0).value
            if type(value) == unicode:
                keylist.append(value.encode('utf-8'))
            else:
                end = True
            row += 1
        return keylist
    def load_list(self,list_file):
        if not list_file:
            keylist = []
        elif list_file[-5:] == '.xlsx':
            keylist = self.__load_xlsx_list(list_file)
        else:
            keylist = open(list_file,'r').read().split('\n')[:-1]
        return keylist
    def __filter(self,row,key_list):
        for value in row:
            for key in key_list:
                if type(key) == str and type(value) == str:
                    if value.find(key) >= 0:
                        return True
        return False
    def __host_filter(self,row):
        key_list = self.hostlist
        if key_list == []:
            return True
        if '序号' in row:
            return True
        else:
            return self.__filter(row,key_list)
    def __black_filter(self,row):
        key_list = self.blacklist
        if key_list == []:
            return False
        else:
            return self.__filter(row,key_list)
    def __table2html(self,table):
        column_widths = []
        PER = 10
        MIN_WIDTH = 2
        MAX_WIDTH = 50
        for row in table:
            for i, value in enumerate(row):
                if type(value) == str:
                    l = len(value.decode('utf-8'))
                    if l == len(value):
                        l = l/2 + 1
                elif type(value) == int:
                    l = len(str(value))
                else:
                    l = len(value)
                if len(column_widths) > i:
                    if l > column_widths[i]:
                            column_widths[i] = l
                else:
                    column_widths.append(l)
        total_width = 0
        for width in column_widths:
            if width < MAX_WIDTH:
                total_width += width
            else:
                total_width += MAX_WIDTH
        tablestyle = "width=%s style='BORDER-BOTTOM-STYLE: solid; BORDER-RIGHT-STYLE: solid; BORDER-TOP-STYLE: solid; BORDER-LEFT-STYLE: solid' border=1 cellSpacing=0 borderColor=#000000 cellPadding=1 bgColor=#ffffff"%(total_width*PER)
        tdstyle = 'style="white-space:nowrap"'
        html = '<table %s>\n'%tablestyle
        for row in table:
            if row[0] == '序号':
                html += '  <tr>\n'
                for i, val in enumerate(row):
                    if column_widths[i] < MIN_WIDTH:
                        width = MIN_WIDTH
                    elif column_widths[i] > MAX_WIDTH:
                        width = MAX_WIDTH
                    else:
                        width = column_widths[i]
                    html += "<td width=%s ><NOBR><FONT size=2 face=宋体 color=#0909f7><STRONG>%s</STRONG></FONT></NOBR></td>"%(width*PER,val)
                html += '  </tr>\n'
            else:
                html += '  <tr>\n'
                for val in row:
                    html += "<td style='word-break:break-all' ><FONT size=2 face=宋体  >%s</FONT></td>"%val
                html += '  </tr>\n'
        html += '</table>'
        return html.decode('utf-8')
    def __xlsx_setstyle(self,style,style_string):
        argv = style_string.split(':')
        n = 0
        fields = {'font':('name','size','bold','italic','superscript','subscript','underline','strikethrough','color'),
                'fill':('fill_type','rotation','start_color','end_color'),
                'borders':('left','right','top','bottom','diagonal','diagonal_direction','all_borders','outline','inside','vertical','horizontal'),
                'alignment':('horizontal','vertical','text_rotation','wrap_text','shrink_to_fit','indent'),
                'number_format':('_format_code','_format_index'),
                'protection':('locked','hidden')
                }
        for obj in ('font','fill','borders','alignment','number_format','protection'):
            for val in fields[obj]:
                if obj == 'borders' and val != 'diagonal_direction':
                    for val_1 in ('border_style','color'):
                        exec "style.%s.%s.%s = eval(argv[n])"%(obj,val,val_1)
                        n += 1
                else:
                    exec "style.%s.%s = eval(argv[n])"%(obj,val)
                    n += 1
        return style
    def get_html(self,titles = 'all'):
        tables = self.tables
        output = ''
        if titles == 'all':
            for title in self.titles:
                output += ('<p><font color="#38f709" size="4">%s</font></p>\n'%title)
                output += self.__table2html(tables[title])
                output += '<br/>\n'
        else:
            for title in titles.split(','):
                try:
                    title = title.decode('utf-8')
                except:
                    title = title.decode('gb18030')
                output += ('<p>%s</p>\n'%title)
                output += self.__table2html(tables[title])
                output += '<br/>\n'
        return output.encode('utf-8')
    def get_xlsx(self,dest_filename):
        tables = self.tables
        wb = Workbook()
        ws = wb.worksheets[0]
        row_num = 0
        column_widths = []
        for title in self.titles:
#            ws.title = title
            row_num += 1
            cell = ws.cell('%s%s'%(get_column_letter(1), row_num))
            cell.value = title
            cell.style.font.size = 13
            row_num += 1
            for row in tables[title]:
                row_num += 1
                col_num = 0
                th = False
                for i, value in enumerate(row):
                    # count column width
                    if type(value) == str:
                        l = len(value.decode('utf-8'))
                        if l == len(value):
                            pass
                        else:
                            l = l*2
                    elif type(value) == int:
                        l = len(str(value))
                    else:
                        l = len(value)
                    if len(column_widths) > i:
                        if l > column_widths[i]:
                                column_widths[i] = l
                    else:
                        column_widths.append(l)

                    col_num += 1
                    cell = ws.cell('%s%s'%(get_column_letter(col_num), row_num))
                    cell.value = value
                    cell.style.font.size = 10
                    for key in ('left','right','top','bottom'):
                        exec "cell.style.borders.%s.border_style = 'medium'"%key
                    if value == '序号':
                        th = True
                    if th:
#                        self.__xlsx_setstyle(cell.style,self.XLSXSTYLE_title)
                        cell.style.font.bold = True
                    else:
                        cell.style.alignment.wrap_text = True
            row_num += 1
        for i, column_width in enumerate(column_widths):
            if column_width < 6:
                ws.column_dimensions[get_column_letter(i+1)].width = 6
            elif column_width > 40:
                ws.column_dimensions[get_column_letter(i+1)].width = 40
            else:
                ws.column_dimensions[get_column_letter(i+1)].width = column_width

#            ws = wb.create_sheet()
#        wb.worksheets.pop()

        wb.save(filename = dest_filename)

def get_html(filename,titles = 'all',host_file = None,blacklist = None):
    t = Tables()
    if host_file:
        t.hostlist = t.load_list(host_file)
    if blacklist:
        t.blacklist = t.load_list(blacklist)
    t.load(filename)
    return t.get_html(titles)

if __name__ == "__main__":
    if len(sys.argv) == 5:
        print get_html(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4])
    elif len(sys.argv) == 4:
        print get_html(sys.argv[1],sys.argv[2],sys.argv[3])
    elif len(sys.argv) == 3:
        print get_html(sys.argv[1],sys.argv[2])
    else:
        print """Usage:
    readxlsx.py <FILE> all/title1[,title2] [HOST_FILE] [Blacklist_FILE]

FILE can be an unix text file or xlsx file.
        """
