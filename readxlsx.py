#!/usr/bin/env python2
# -*- coding: utf-8 -*-
import os,sys
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import openpyxl.style
from bs4 import BeautifulSoup

class Tables():
    XLSXSTYLE = "'Calibri':10:False:False:False:False:'none':False:'FF000000':'none':0:'FFFFFFFF':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'none':'FF000000':0:'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'general':'center':0:True:False:0:'General':0:'inherit':'inherit'"
    XLSXSTYLE_title = "'Calibri':10:True:False:False:False:'none':False:'FF0000FF':'none':0:'FFFFFFFF':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'medium':'FF000000':'none':'FF000000':0:'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'none':'FF000000':'general':'center':0:False:False:0:'General':0:'inherit':'inherit'"
    def __init__(self):
        self.titles = []
        self.tables = {}
        self.blocklist = []
    def load(self,filename):
        ext = filename[filename.find('.')+1:]
        if ext == "xlsx":
            self.load_xlsx(filename)
        elif ext in ('htm','html'):
            self.load_html(filename)
        else:
            print "error: can load '%s' file!"%ext
    def load_html(self,html_file):
        html = file(os.path.realpath(html_file),'r').read().decode('utf-8')
        soup = BeautifulSoup(html)
        tables = soup.find_all('table')
        for table in tables:
            title = table.find_previous('p').get_text()
            self.titles.append(title)
            self.tables[title] = []
            for row in table.find_all('tr'):
                line = []
                for val in row.find_all('td'):
                    line.append(val.get_text())
                self.tables[title].append(line)

    def load_xlsx(self,xlsx_file):
        wb = load_workbook(filename = xlsx_file)
        sheet_ranges = wb.get_active_sheet()
        table = 0
        row = 0
        row_empty = 0
        row_max_empty = 100
        end = False
        th = None
        while not end:
            col = []
            col_empty = 0
            col_max_empty = 3
            col_end = False
            while not col_end:
                value = sheet_ranges.cell(row=row,column=len(col)).value
                if value:
                    col_empty = 0
                else:
                    value = ''
                    col_empty += 1
                col.append(value)
                if col_empty >= col_max_empty:
                    col = col[:-col_max_empty]
                    col_end = True
            if len(col) == 1:
                if not th:
                    th = col[0]
                    worktable = []
                else:
                    if th == "" and worktable == []:
                        pass
                    else:
                        self.titles.append(th)
                        self.tables[th] = worktable
                    th = col[0]
                    worktable = []
            elif len(col) > 1:
                if not th:
                    th = ""
                    worktable = []
                worktable.append(col)
                row_empty = 0
            else:
                row_empty += 1
            if row_empty >= row_max_empty:
                end = True
                self.titles.append(th)
                self.tables[th] = worktable
            row += 1
        return 0
    def __load_xlsx_hostlist(self,xlsx_file):
        wb = load_workbook(filename = xlsx_file)
        sheet_ranges = wb.get_active_sheet()
        end = False
        row = 0
        hostlist = []
        while not end:
            value = sheet_ranges.cell(row=row,column=0).value
            if type(value) == unicode:
                hostlist.append(value.encode('utf-8'))
            else:
                end = True
            row += 1
        return hostlist
    def __filter(self,hostlist_file):
        if not hostlist_file:
            return self.tables
        elif hostlist_file[-5:] == '.xlsx':
            hostlist = self.__load_xlsx_hostlist(hostlist_file)
        else:
            hostlist = open(hostlist_file,'r').read().split('\n')[:-1]
        tables = {}
        for title in self.titles:
            table_new = []
            for row in self.tables[title]:
                if u'序号' in row:
                    table_new.append(row)
                else:
                    result = False
                    drop = False
                    n = 0
                    while not result:
                        for host in hostlist:
                            if result == False and type(host) == unicode and type(row[n]) == unicode:
                                if row[n].find(host.decode('utf-8')) >= 0:
                                    result = True
                                    for col in row:
                                        if type(col) == unicode:
                                            for val in self.blocklist:
                                                if col.find(val.decode('utf-8')) >= 0:
                                                    drop = True
                                    if drop:
                                        pass
                                    else:
                                        table_new.append(row)
                        if n >= len(row) - 1:
                            result = True
                        n += 1
            tables[title] = table_new
        return tables
    def __table2html(self,table):
        WIDTH = {'序号':40,
                }
        tablestyle = "width=1300 style='BORDER-BOTTOM-STYLE: solid; BORDER-RIGHT-STYLE: solid; BORDER-TOP-STYLE: solid; BORDER-LEFT-STYLE: solid' border=1 cellSpacing=0 borderColor=#000000 cellPadding=1 bgColor=#ffffff"
        tdstyle = 'style="white-space:nowrap"'
        html = '<table %s>\n'%tablestyle
        for row in table:
            if row[0] == u'序号':
                html += '  <tr>\n'
                for val in row:
                    if WIDTH.has_key(val.encode('utf-8')):
                        width = WIDTH[val.encode('utf-8')]
                        html += u"<td width=%s ><NOBR><FONT size=2 face=宋体 color=#0909f7><STRONG>%s</STRONG></FONT></NOBR></td>"%(width,val)
                    else:
                        html += u"<td width=120 ><NOBR><FONT size=2 face=宋体 color=#0909f7><STRONG>%s</STRONG></FONT></NOBR></td>"%val
                html += '  </tr>\n'
            else:
                html += '  <tr>\n'
                for val in row:
                    html += u"<td style='word-break:break-all' ><FONT size=2 face=宋体  >%s</FONT></td>"%val
                html += '  </tr>\n'
        html += '</table>'
        return html
    def __xlsx_setstyle(self,style,style_string):
        argv = style_string.split(':')
        n = 0
        fields = {'font':('name','size','bold','italic','superscript','subscript','underline','strikethrough','color'),
                'fill':('fill_type','rotation','start_color','end_color'),
                'borders':('left','right','top','bottom','diagonal','diagonal_direction','all_borders','outline','inside','vertical','horizontal'),
                'alignment':('horizontal','vertical','text_rotation','wrap_text','shrink_to_fit','indent'),
                'number_format':('_format_code','_format_index'),
                'protection':('locked','hidden')
                }
        for obj in ('font','fill','borders','alignment','number_format','protection'):
            for val in fields[obj]:
                if obj == 'borders' and val != 'diagonal_direction':
                    for val_1 in ('border_style','color'):
                        exec "style.%s.%s.%s = eval(argv[n])"%(obj,val,val_1)
                        n += 1
                else:
                    exec "style.%s.%s = eval(argv[n])"%(obj,val)
                    n += 1
        return style
    def get_html(self,titles = 'all',host_file = None):
        tables = self.__filter(host_file)
        output = ''
        if titles == 'all':
            for title in self.titles:
                output += ('<p>%s</p>\n'%title)
                output += self.__table2html(tables[title])
                output += '<br/>\n'
        else:
            for title in titles.split(','):
                try:
                    title = title.decode('utf-8')
                except:
                    title = title.decode('gb18030')
                output += ('<p>%s</p>\n'%title)
                output += self.__table2html(tables[title])
                output += '<br/>\n'
        return output.encode('utf-8')
    def get_xlsx(self,dest_filename,host_file = None):
        tables = self.__filter(host_file)
        wb = Workbook()
        ws = wb.worksheets[0]
        column_widths = []
        for title in self.titles:
            ws.title = title
            row_num = 0
            for row in tables[title]:
                row_num += 1
                col_num = 0
                th = False
                for i, value in enumerate(row):
                    col_num += 1
                    cell = ws.cell('%s%s'%(get_column_letter(col_num), row_num))
                    cell.value = value
                    if value == u'序号':
                        th = True
                    if th:
                        self.__xlsx_setstyle(cell.style,self.XLSXSTYLE_title)
                    else:
                        self.__xlsx_setstyle(cell.style,self.XLSXSTYLE)
                    # count column width
                    if len(column_widths) > i:
                        if type(value) == unicode:
                            if len(value) > column_widths[i]:
                                column_widths[i] = len(value)
                        elif type(value) == int:
                            if len(str(value)) > column_widths[i]:
                                column_widths[i] = len(str(value))
                        else:
                            if 5 > column_widths[i]:
                                column_widths[i] = 5
                    else:
                        if type(value) == unicode:
                            column_widths.append(len(value))
                        elif type(value) == int:
                            column_widths.append(len(str(value)))
                        else:
                            column_widths.append(5)

            for i, column_width in enumerate(column_widths):
                if column_width < 5:
                    ws.column_dimensions[get_column_letter(i+1)].width = 6
                elif column_width > 40:
                    ws.column_dimensions[get_column_letter(i+1)].width = 30
                else:
                    ws.column_dimensions[get_column_letter(i+1)].width = column_width
            ws = wb.create_sheet()
        wb.worksheets.pop()

        wb.save(filename = dest_filename)

def get_html(filename,titles = 'all',host_file = None,blacklist = []):
    t = Tables()
    t.load(filename)
    t.blacklist = blacklist
    t.get_xlsx("%s.xlsx"%filename)
#    return t.get_html(titles,host_file)

if __name__ == "__main__":
    if len(sys.argv) == 4:
        print get_html(sys.argv[1],sys.argv[2],sys.argv[3])
    elif len(sys.argv) == 3:
        print get_html(sys.argv[1],sys.argv[2])
    else:
        print """Usage:
    readxlsx.py <name>.xlsx all/title1[,title2] [HOSTFILE]

HOSTFILE can be an unix text file or xlsx file.
        """
